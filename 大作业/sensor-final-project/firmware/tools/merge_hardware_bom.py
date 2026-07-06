from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
HARDWARE = ROOT / "hardware"
EDA_BOM = HARDWARE / "BOM.csv"
RAW_EDA_BOM = HARDWARE / "BOM_eda_raw.csv"
COST_BOM = HARDWARE / "BOM.xlsx"
OUT_CSV = HARDWARE / "BOM.csv"
OUT_XLSX = HARDWARE / "BOM_merged.xlsx"


OUT_COLUMNS = [
    "section",
    "item",
    "quantity",
    "designator",
    "footprint",
    "value",
    "manufacturer_part",
    "manufacturer",
    "supplier_part",
    "supplier",
    "unit_price_cny",
    "subtotal_cny",
    "note",
    "source",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def number(value: Any) -> Any:
    if value in (None, ""):
        return ""
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return value


def read_text_auto(path: Path) -> str:
    data = path.read_bytes()
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        return data.decode("utf-16")
    for encoding in ("utf-8-sig", "gbk"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def get_eda_source() -> Path:
    if RAW_EDA_BOM.exists():
        return RAW_EDA_BOM

    raw = read_text_auto(EDA_BOM)
    first_line = raw.splitlines()[0] if raw.splitlines() else ""
    if "Comment" not in first_line or "Designator" not in first_line:
        raise RuntimeError(
            "BOM.csv does not look like the original EDA export, and BOM_eda_raw.csv was not found."
        )

    RAW_EDA_BOM.write_text(raw, encoding="utf-8-sig")
    return RAW_EDA_BOM


def read_eda_bom(path: Path) -> list[dict[str, Any]]:
    raw = read_text_auto(path)
    first_line = raw.splitlines()[0] if raw.splitlines() else ""
    delimiter = "\t" if "\t" in first_line else ","
    rows: list[dict[str, Any]] = []
    for row in csv.DictReader(raw.splitlines(), delimiter=delimiter):
        if not any(text(v) for v in row.values()):
            continue
        item = text(row.get("Comment")) or text(row.get("Manufacturer Part"))
        rows.append(
            {
                "section": "EDA原始BOM",
                "item": item,
                "quantity": number(row.get("Quantity")),
                "designator": text(row.get("Designator")),
                "footprint": text(row.get("Footprint")),
                "value": text(row.get("Value")),
                "manufacturer_part": text(row.get("Manufacturer Part")),
                "manufacturer": text(row.get("Manufacturer")),
                "supplier_part": text(row.get("Supplier Part")),
                "supplier": text(row.get("Supplier")),
                "unit_price_cny": "",
                "subtotal_cny": "",
                "note": "来自EDA导出的器件/封装/位号信息",
                "source": path.name,
            }
        )
    return rows


def read_cost_bom(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for cells in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: cells[i] if i < len(cells) else "" for i in range(len(headers))}
        item = text(row.get("item"))
        if not item:
            continue
        rows.append(
            {
                "section": "成本补充",
                "item": item,
                "quantity": number(row.get("quantity")),
                "designator": "",
                "footprint": "",
                "value": "",
                "manufacturer_part": "",
                "manufacturer": "",
                "supplier_part": "",
                "supplier": "",
                "unit_price_cny": number(row.get("unit_price")),
                "subtotal_cny": number(row.get("subtotal")),
                "note": text(row.get("note")),
                "source": path.name,
            }
        )
    return rows


def is_total(row: dict[str, Any]) -> bool:
    return "总计" in row["item"] or row["item"].lower() in {"total", "sum"}


def key_item(name: str) -> str:
    upper = name.upper()
    for key in ["MPU6050", "HMC5883L", "GY-273", "BMP280", "ESP32", "GPS", "LED"]:
        if key in upper:
            return key
    if "排针" in name or "连接器" in name or "PZ254" in upper:
        return "CONNECTOR"
    if "阻容" in name or "1K" in upper or "1KΩ" in upper:
        return "PASSIVE"
    if "PCB" in upper:
        return "PCB"
    if "焊接" in name or "辅料" in name:
        return "ASSEMBLY"
    return upper


def merge_rows(eda_rows: list[dict[str, Any]], cost_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cost_by_key: dict[str, dict[str, Any]] = {}
    total_rows: list[dict[str, Any]] = []
    for row in cost_rows:
        if is_total(row):
            total_rows.append(row)
            continue
        cost_by_key.setdefault(key_item(row["item"]), row)

    used_cost_keys: set[str] = set()
    merged: list[dict[str, Any]] = []
    for row in eda_rows:
        item_key = key_item(row["item"])
        cost = cost_by_key.get(item_key)
        out = dict(row)
        if cost:
            used_cost_keys.add(item_key)
            out["unit_price_cny"] = cost["unit_price_cny"]
            out["subtotal_cny"] = cost["subtotal_cny"]
            out["note"] = f"{row['note']}；成本信息：{cost['item']}，{cost['note']}".strip("；")
            out["source"] = f"{row['source']} + {cost['source']}"
        else:
            out["note"] = f"{row['note']}；成本表未填写该项单价"
        merged.append(out)

    for row in cost_rows:
        item_key = key_item(row["item"])
        if is_total(row):
            continue
        if item_key not in used_cost_keys:
            merged.append(row)

    merged.extend(total_rows)
    return merged


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in OUT_COLUMNS})


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM_merged"
    ws.append(OUT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in OUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 14,
        "B": 28,
        "C": 10,
        "D": 18,
        "E": 28,
        "F": 12,
        "G": 20,
        "H": 20,
        "I": 16,
        "J": 12,
        "K": 14,
        "L": 14,
        "M": 42,
        "N": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    eda_rows = read_eda_bom(get_eda_source())
    cost_rows = read_cost_bom(COST_BOM)
    merged = merge_rows(eda_rows, cost_rows)
    write_csv(OUT_CSV, merged)
    write_xlsx(OUT_XLSX, merged)
    print(f"merged_rows={len(merged)}")
    print(f"csv={OUT_CSV}")
    print(f"xlsx={OUT_XLSX}")


if __name__ == "__main__":
    main()
